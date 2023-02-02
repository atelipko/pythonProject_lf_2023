import openpyxl

def example(list_):
    """
    Создание xlsx файла и запись в него
    """
    # создаю книгу
    book = openpyxl.Workbook()

    # по умолчанию создается с таблицей Sheet
    # print(book.sheetnames)
    book.remove(book.active)




    # создаю таблицы
    # book.active.title = "Коллеги"
    sheet = book.create_sheet("Dnipro",0) # таблица будет первой

    """sheet_Ilarionove = book.create_sheet("Ilarionove")
    sheet_Kamyanske = book.create_sheet("Kamyanske")
    sheet_Pavlograd = book.create_sheet("Pavlograd")
    sheet_Novomoskovsk = book.create_sheet("Novomoskovsk")
    sheet_Synelnykove = book.create_sheet("Synelnykove")
    sheet_Novomoskovsk = book.create_sheet("Novomoskovsk")
    """
    caunt=1
    for i in range(len(list_)):
        for j in range(len(list_[i].my_list_Finih)):
            A=list_[i].my_list_Finih[j][0] + " - " + list_[i].my_list_Finih[j][1] + " = " + str(list_[i].my_list_Finih[j][2])+" m"
            B=list_[i].my_list_Finih[j][0]
            C=list_[i].my_list_Finih[j][1]
            D=list_[i].my_list_Finih[j][2]
            sheet['A' + str(caunt)] = A
            sheet['B' + str(caunt)] = B
            sheet['C' + str(caunt)] = C
            sheet['D' + str(caunt)] = D
            caunt=caunt+1

    """for sheet in book.worksheets:  # перебираю таблицы
        for row in data_samples():  # получаю данные
            sheet.append(row)  # записываю данные в строки таблиц"""

    book.save("test.xlsx")